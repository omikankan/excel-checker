# -*- coding: utf-8 -*-

import tkinter.messagebox
import tkinter.filedialog
import os
from openpyxl import load_workbook


class FileAttributes:
    def __init__(self):
        self.definedNames = ""
        self.sheetsName = ""
        self.styles = ""


class ExcelData:
    def __init__(self, path):
        self.path = path
        self.fileAttributes = FileAttributes()

    def get_attributes(self):
        book = load_workbook(self.path)

        self.fileAttributes.definedNames = ""
        for name in book.defined_names.definedName:
            if self.fileAttributes.definedNames != "" \
                    : self.fileAttributes.definedNames = self.fileAttributes.definedNames + "\n"
            hidden = ""
            if name.hidden is not None : hidden = " ←hidden"
            self.fileAttributes.definedNames = self.fileAttributes.definedNames + "  " + name.name + hidden

        self.fileAttributes.sheetsName = ""
        for sheetname in book.sheetnames:
            if self.fileAttributes.sheetsName != "" \
                    : self.fileAttributes.sheetsName = self.fileAttributes.sheetsName + "\n"
            hidden = ""
            if book[sheetname].sheet_state == "hidden" : hidden = " ←hidden"
            self.fileAttributes.sheetsName = self.fileAttributes.sheetsName + "  " + sheetname + hidden

        self.fileAttributes.styles = ""
        for style in book.style_names:
            if self.fileAttributes.styles != "" : self.fileAttributes.styles = self.fileAttributes.styles + "\n"
            self.fileAttributes.styles = self.fileAttributes.styles + "  " + style

        return self.fileAttributes


class FileOpenDialog(tkinter.Frame):
    filetype = [('', "*.xls"), ('', "*.xlsx"), ('', "*.xlsm")]
    path = ""

    def __init__(self, master):
        super().__init__(master)
        master.withdraw()
        self.path = tkinter.filedialog.askopenfilename(filetypes=self.filetype)

    def open(self):
        attributes = None
        if os.path.isfile(self.path) : attributes = ExcelData(self.path).get_attributes()
        return attributes


class Output:
    def __init__(self, file_attributes):
        self.fileAttributes = file_attributes

    def out(self):
        print("##### result #####")
        print("definedNames : \n" + self.fileAttributes.definedNames)
        print("styles : \n" + self.fileAttributes.styles)
        print("sheetsName : \n" + self.fileAttributes.sheetsName)


### main ###
root = tkinter.Tk()
fileOpenDialog = FileOpenDialog(root)
root.attributes = fileOpenDialog.open()

if root.attributes is not None:
    outputDialog = Output(root.attributes)
    outputDialog.out()
